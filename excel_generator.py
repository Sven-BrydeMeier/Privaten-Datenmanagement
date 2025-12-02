"""
Excel-Generator für RHM Posteingang
Erstellt Excel-Dateien pro Sachbearbeiter und Gesamt-Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, List
from pathlib import Path


class ExcelGenerator:
    SPALTEN = [
        'Eingangsdatum',          # A
        'Internes Aktenzeichen',  # B
        'Externes Aktenzeichen',  # C
        'Mandant',                # D
        'Gegner / Absender',      # E
        'Absendertyp',            # F
        'Sachbearbeiter',         # G
        'Fristdatum',             # H
        '⚠',                      # I - Frist-Indikator (Roter Kreis)
        'Fristtyp',               # J
        'Fristquelle',            # K
        'Textauszug',             # L
        'PDF-Datei',              # M
        'Status'                  # N
    ]

    def __init__(self):
        self.heute = datetime.now().date()

    def erstelle_excel_dateien(self, alle_daten: List[Dict], temp_path: Path) -> Dict[str, bytes]:
        """
        Erstellt Excel-Dateien pro Sachbearbeiter

        Returns:
            Dict mit Sachbearbeiter → Excel-Bytes
        """
        excel_dateien = {}

        # Gruppiere nach Sachbearbeiter
        sb_gruppen = {}
        for daten in alle_daten:
            sb = daten['sachbearbeiter']
            if sb not in sb_gruppen:
                sb_gruppen[sb] = []
            sb_gruppen[sb].append(daten)

        # Erstelle Excel pro Sachbearbeiter
        for sb, daten_liste in sb_gruppen.items():
            excel_bytes = self._erstelle_einzelne_excel(daten_liste, sb)
            excel_dateien[sb] = excel_bytes

        return excel_dateien

    def erstelle_gesamt_excel(self, alle_daten: List[Dict]) -> bytes:
        """
        Erstellt Gesamt-Excel mit allen Dokumenten
        """
        return self._erstelle_einzelne_excel(alle_daten, "Gesamt")

    def _erstelle_einzelne_excel(self, daten_liste: List[Dict], titel: str) -> bytes:
        """
        Erstellt eine einzelne Excel-Datei
        """
        zeilen = []

        for daten in daten_liste:
            akt_info = daten['aktenzeichen_info']
            analyse = daten['analyse']
            sb = daten['sachbearbeiter']
            dateiname = daten['dateiname']

            # Fristen verarbeiten
            fristen = analyse.get('fristen', [])

            if fristen:
                # Eine Zeile pro Frist
                for frist in fristen:
                    zeile = self._erstelle_zeile(
                        akt_info, analyse, sb, dateiname, frist
                    )
                    zeilen.append(zeile)
            else:
                # Eine Zeile ohne Frist
                zeile = self._erstelle_zeile(
                    akt_info, analyse, sb, dateiname, None
                )
                zeilen.append(zeile)

        # DataFrame erstellen
        df = pd.DataFrame(zeilen, columns=self.SPALTEN)

        # Excel in BytesIO schreiben
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Fristen', index=False)

        buffer.seek(0)
        excel_bytes = buffer.getvalue()

        # Farbliche Markierung hinzufügen
        excel_bytes = self._markiere_fristen(excel_bytes)

        return excel_bytes

    def _erstelle_zeile(self, akt_info: Dict, analyse: Dict, sb: str,
                       dateiname: str, frist: Dict = None) -> List:
        """
        Erstellt eine Zeile für die Excel-Tabelle
        """
        # Deutsches Datumsformat: DD.MM.YYYY
        heute_str = self.heute.strftime('%d.%m.%Y')

        # Frist-Indikator: Roter Kreis wenn Frist vorhanden
        frist_indikator = '🔴' if frist else ''

        # Fristdatum in deutsches Format konvertieren
        fristdatum_str = ''
        if frist and frist.get('datum'):
            try:
                # Parse Datum (Format von KI: YYYY-MM-DD)
                frist_datum = datetime.strptime(frist.get('datum'), '%Y-%m-%d')
                # Konvertiere zu deutschem Format
                fristdatum_str = frist_datum.strftime('%d.%m.%Y')
            except:
                # Falls Parsing fehlschlägt, verwende Original
                fristdatum_str = frist.get('datum', '')

        zeile = [
            heute_str,  # Eingangsdatum (deutsches Format)
            akt_info.get('internes_az', ''),  # Internes AZ
            ', '.join(akt_info.get('externe_az', [])),  # Externe AZ
            analyse.get('mandant', ''),  # Mandant
            analyse.get('gegner', ''),  # Gegner / Absender
            analyse.get('absendertyp', 'Sonstige'),  # Absendertyp
            sb,  # Sachbearbeiter
            fristdatum_str,  # Fristdatum (deutsches Format)
            frist_indikator,  # ⚠ Frist-Indikator
            frist.get('typ', '') if frist else '',  # Fristtyp
            frist.get('quelle', '') if frist else '',  # Fristquelle
            analyse.get('textauszug', '')[:200],  # Textauszug (gekürzt)
            dateiname,  # PDF-Datei
            'Neu'  # Status
        ]

        return zeile

    def _markiere_fristen(self, excel_bytes: bytes) -> bytes:
        """
        Professionelle Excel-Formatierung:
        - Farbige Kopfzeilen (dunkelblau mit weißer Schrift)
        - Zebra-Streifen für bessere Lesbarkeit
        - Borders um alle Zellen
        - Frist-Highlighting:
          * Rot: ≤ 3 Tage
          * Orange: ≤ 7 Tage
          * Gelb: ≤ 14 Tage
        - Automatische Spaltenbreite
        """
        buffer = BytesIO(excel_bytes)
        wb = load_workbook(buffer)
        ws = wb['Fristen']

        # === FARBEN DEFINIEREN ===
        # Header
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dunkelblau
        header_font = Font(bold=True, color="FFFFFF", size=11)  # Weiß, Fett

        # Fristen (ganze Zeile)
        frist_rot = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rot
        frist_orange = PatternFill(start_color="FFB84D", end_color="FFB84D", fill_type="solid")  # Orange
        frist_gelb = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Gelb

        # Zebra-Streifen
        zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Hellgrau

        # Borders
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Alignment
        center_aligned = Alignment(horizontal='center', vertical='center')
        left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # === SPALTENBREITEN SETZEN ===
        column_widths = {
            'A': 14,   # Eingangsdatum
            'B': 18,   # Internes AZ
            'C': 20,   # Externes AZ
            'D': 20,   # Mandant
            'E': 20,   # Gegner
            'F': 15,   # Absendertyp
            'G': 12,   # Sachbearbeiter
            'H': 12,   # Fristdatum
            'I': 5,    # ⚠ Frist-Indikator
            'J': 18,   # Fristtyp
            'K': 35,   # Fristquelle
            'L': 40,   # Textauszug
            'M': 35,   # PDF-Datei
            'N': 10    # Status
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # === HEADER-ROW FORMATIEREN ===
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = thin_border

        # Spalte H = Fristdatum (Index 8)
        fristdatum_col = 8

        # === DATEN-ROWS FORMATIEREN ===
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fristdatum_cell = row[fristdatum_col - 1]  # -1 wegen 0-Index
            fristdatum_value = fristdatum_cell.value

            # Prüfe Frist-Dringlichkeit
            frist_fill = None
            if fristdatum_value:
                try:
                    # Parse Datum (deutsches Format: DD.MM.YYYY)
                    if isinstance(fristdatum_value, str):
                        # Versuche deutsches Format
                        try:
                            frist_datum = datetime.strptime(fristdatum_value, '%d.%m.%Y').date()
                        except ValueError:
                            # Fallback: ISO-Format (für Kompatibilität)
                            frist_datum = datetime.strptime(fristdatum_value, '%Y-%m-%d').date()
                    elif isinstance(fristdatum_value, datetime):
                        frist_datum = fristdatum_value.date()
                    else:
                        frist_datum = None

                    if frist_datum:
                        # Berechne Differenz
                        diff = (frist_datum - self.heute).days

                        # Wähle Farbe
                        if diff <= 3:
                            frist_fill = frist_rot
                        elif diff <= 7:
                            frist_fill = frist_orange
                        elif diff <= 14:
                            frist_fill = frist_gelb

                except Exception as e:
                    # Bei Fehler: Keine Farb-Hervorhebung
                    pass

            # Formatiere alle Zellen in der Zeile
            for col_idx, cell in enumerate(row, start=1):
                # Border
                cell.border = thin_border

                # Frist-Highlighting (ganze Zeile)
                if frist_fill:
                    cell.fill = frist_fill
                # Zebra-Streifen (nur wenn keine Frist-Hervorhebung)
                elif row_idx % 2 == 0:
                    cell.fill = zebra_fill

                # Alignment
                if col_idx in [1, 2, 7, 8, 9, 14]:  # Datum, AZ, SB, Frist, Indikator, Status
                    cell.alignment = center_aligned
                else:
                    cell.alignment = left_aligned

                # Frist-Indikator-Spalte (I): Größere Schrift für Emoji
                if col_idx == 9 and cell.value:  # Spalte I = Frist-Indikator
                    cell.font = Font(size=14)
                    cell.alignment = center_aligned

        # Zeilen-Höhe anpassen
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 25

        # Header-Zeile etwas höher
        ws.row_dimensions[1].height = 30

        # Freeze erste Zeile (Header)
        ws.freeze_panes = 'A2'

        # Zurück in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
